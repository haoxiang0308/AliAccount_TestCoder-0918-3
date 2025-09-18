import openpyxl
from openpyxl import Workbook
import os

# Create a new workbook
wb = Workbook()
ws = wb.active

# Add some sample data for demonstration
ws['A1'] = 'Product'
ws['B1'] = 'Quantity'
ws['C1'] = 'Date'
ws['D1'] = 'Amount'

sample_data = [
    ['Apples', 10, '2023-01-01', 15.50],
    ['Oranges', 20, '2023-01-02', 30.00],
    ['Bananas', 15, '2023-01-03', 12.75],
    ['Grapes', 5, '2023-01-04', 22.00],
]

for row in sample_data:
    ws.append(row)

# Save the workbook temporarily without macro
temp_filename = 'temp_workbook.xlsx'
wb.save(temp_filename)

# Re-open the workbook as an openpyxl Workbook with macro support
wb_with_macro = openpyxl.load_workbook(temp_filename, keep_vba=True)

# Add the VBA macro
vba_code = """
Sub FormatSheet()
    Dim ws As Worksheet
    Set ws = ActiveSheet

    ' Format the header row
    With ws.Range("A1:D1")
        .Font.Bold = True
        .Font.Size = 12
        .Interior.Color = RGB(200, 200, 200)
        .Borders.LineStyle = xlContinuous
    End With

    ' Auto-fit columns
    ws.Columns("A:D").AutoFit

    ' Add a title
    ws.Range("A1").Value = "Sales Data"
    ws.Range("A1:D1").Merge
    ws.Range("A1").HorizontalAlignment = xlCenter

    ' Format data rows
    Dim lastRow As Long
    lastRow = ws.Cells(ws.Rows.Count, "A").End(xlUp).Row

    Dim i As Long
    For i = 2 To lastRow
        If i Mod 2 = 0 Then
            ws.Range("A" & i & ":D" & i).Interior.Color = RGB(240, 240, 240)
        End If
        ws.Range("A" & i & ":D" & i).Borders.LineStyle = xlContinuous
    Next i

    ' Apply number formatting to the "Amount" column (D)
    ws.Range("D2:D" & lastRow).NumberFormat = "$#,##0.00"

    MsgBox "Sheet formatting complete!"
End Sub
"""

wb_with_macro.vba_code = vba_code

# Save the final macro-enabled workbook
final_filename = 'formatted_sheet.xlsm'
wb_with_macro.save(final_filename)

# Clean up the temporary file
os.remove(temp_filename)

print(f"Macro-enabled workbook '{final_filename}' has been created.")